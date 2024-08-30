import time
import warnings
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.edge.service import Service as EdgeService
import openpyxl as op
# quitar advertencias
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)

# maximizar navegador
options = Options()
options.add_argument("--start-maximized")
edge_service = EdgeService(executable_path="msedgedriver.exe")

driver = webdriver.Edge(service=edge_service, options=options)
#version de sele 4.9.0

excel = op.load_workbook("Ruc.xlsx")
hoja_completar = excel.worksheets[0]
hoja_intrucciones = excel.worksheets[1]
PATH_FILA = "/html/body/div/main/section/div/div[2]/article/div/div[2]/div[2]/div[2]/div/div["
PATH_RUC = "/html/body/div/main/section/div/div[2]/article/div/div[2]/div[2]/div[2]/div/div[1]/div["
"""3]/p"""
PATH_INFO = "/html/body/div[2]/div[3]/header/div/div/div/button[2]/span[1]"
PATH_CORREO= "/html/body/div[2]/div[3]/div[2]/div[2]/div[3]/div/div/div/div/div[1]/span"
PATH_CERRAR = "/html/body/div[2]/div[3]/div[3]/div[1]/button[1]"

CONTADOR_FILA_COMPLETAR = 2
CONTADOR_FILA_INSTRUCCIONES = 1

cantidad_filas = hoja_intrucciones.cell(row=8, column=2).value
cantidad_pestanas =  instruccion1 = hoja_intrucciones.cell(row=9, column=2).value
INSTRUCCIONES = []

link = hoja_intrucciones.cell(row=1, column=2).value
driver.get(link)
var = input("Preionar Enter una vez logeado")
driver.get(link)
time.sleep(2)
INSTRUCCIONES.append(['get', link]) # agregar el link que vamos a abrir
for i in range(1, 19):
    for i in range(1, 3):
        ruc = ""
        correo = ""
        try:
            ruc = driver.find_element(By.XPATH, f"{PATH_RUC}{i}]/p").text
            time.sleep(0.5)
            driver.find_element(By.XPATH, f"{PATH_FILA}{i}]").click()
            time.sleep(0.7)
            driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/header/div/div/div/button[2]/span[1]").click()# click en contactos
            time.sleep(0.7)
            correo = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[2]/div[2]/div[3]/div/div/div/div/div[1]/span").text
            time.sleep(0.5)
            #cerrar ventana
            driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[3]/div[1]/button[1]/span[1]").click()
            time.sleep(0.5)
            print(CONTADOR_FILA_COMPLETAR, ruc, correo)
        except Exception as ex:
            pass
        hoja_completar.cell(CONTADOR_FILA_COMPLETAR, 1).value = ruc
        hoja_completar.cell(CONTADOR_FILA_COMPLETAR, 2).value = correo
        CONTADOR_FILA_COMPLETAR += 1
        excel.save(hoja_intrucciones.cell(row=10, column=2).value)
    try:
        driver.find_element(By.XPATH, "/html/body/div/main/section/div/div[2]/article/div/div[3]/div/button[3]").click()
        print("Boton Siguiente")
        time.sleep(2)
    except:
        print("No se clickeo")
    try:
        driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[3]/div[1]/button[1]/span[1]").click()
        time.sleep(0.5)
    except:
        pass


excel.save(hoja_intrucciones.cell(row=10, column=2).value)
driver.close()