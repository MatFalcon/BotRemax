import openpyxl as op
from pathlib import PurePath, Path
import pandas as pd
import numpy


RUTA_BOT = PurePath(Path().absolute())
RUTA_DATOS = PurePath(RUTA_BOT, "datos")
RUTA_DRIVER = f"{PurePath(RUTA_BOT, "driver")}\\msedgedriver.exe"
RUTA_ARCHIVO_CSV = PurePath(RUTA_BOT, 'driver', 'remax_propiedades.csv')
RUTA_EXCEL = PurePath(RUTA_BOT, "driver", "Pass.xlsx")



def obtener_credenciales_excel(hoja):
    dic = {}
    # mas uno para que tome hasta el ultimo neto
    max_fila = hoja.max_row + 1
    for fila in range(2, max_fila):

        dic[hoja.cell(fila, 1).value] = {
            "correo": hoja.cell(fila, 2).value,
            "contra": hoja.cell(fila, 3).value,
            "ingresa": hoja.cell(fila, 4).value,
            "telefono": hoja.cell(fila, 5).value
        }
    return dic


def crenciales_paginas():
    """
        hendy, clasi, info
    """

    excel = op.load_workbook(RUTA_EXCEL, read_only=True, data_only=True)
    hoja_info = excel.worksheets[0]
    hoja_hendy = excel.worksheets[1]
    hoja_clasi = excel.worksheets[2]

    dic = {
    "hendy":obtener_credenciales_excel(hoja_hendy),
    "info":obtener_credenciales_excel(hoja_info),
    "clasi":obtener_credenciales_excel(hoja_clasi)
    }
    excel.close()
    return dic










































