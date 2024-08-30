import os
import ssl
import time
import openpyxl
import warnings
import pandas as pd
import urllib.request

from datetime import datetime
from selenium import webdriver
from pathlib import PurePath, Path
from para_log import escribir_en_log
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions

# Desactivar todas las advertencias de Pandas
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)
options = Options()
options.add_argument("--start-maximized")

hay_resultados = True

# rutas relativas para el bot
RUTA_BOT = PurePath(Path().absolute())
RUTA_DATOS = PurePath(RUTA_BOT, "datos")
RUTA_DF = PurePath(RUTA_BOT, "driver", "remax_propiedades.csv")
print(RUTA_DATOS)
RUTA_DRIVER = f"{PurePath(RUTA_BOT, "driver")}\\msedgedriver.exe"
print(RUTA_DRIVER)
print(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.csv'))
# Obtener la fecha y hora actual


ciudades = ["Asuncion", "Sanber", "Fernando", "Luque", "Sanlo", "VillaElisa"]
fecha_actual = datetime.now()

url = "https://www.remax.com.py/"
# path para inicio
path_campo_ciudad = "/html/body/form/div[3]/div[3]/div[1]/div/div[3]/div/div/div/div[2]/div[2]/input"
path_boton_siguiente = "/html/body/div[1]/form/div[3]/div[5]/div/div[8]/div/div[3]/div/div/div[3]/div[2]/div[1]/div[3]/div[2]/div/nav/ul/li[7]/a"
# para el bucle en el que quita los datos de cada propiedad
tipo_propiedad_excluir = ["RESERVADO", "VENDIDO"]
path_tipo_propiedad = ["/html/body/div[1]/form/div[3]/div[5]/div/div[8]/div/div[3]/div/div/div[3]/div[2]/div[1]/div[1]/div/div[","]/div/div[7]/span"]

# paths para quitar datos por cada propiedad
path_titulo = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div[1]/div[1]/h1"
path_precio = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div[1]/div[3]/div/a"
path_precio2 = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div[1]/div[4]/div/a"
path_direccion = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div[1]/div[5]"
path_id = ["/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div[1]/div[","]"]

path_atributo = ["/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div[2]/div[1]/div[","]/div/div[1]/span"]
path_valor = ["/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div[2]/div[1]/div[","]/div/div[2]/span"]
path_sub_atri = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[1]/div[2]/div[2]/div[2]/div/div[1]"
path_descripcion = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div/div"
path_descripcion2 = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[2]/div[2]/div/div/div[2]/div"
path_descripcion3 = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[3]/div[2]/div/div/div[2]/div/div/div"
path_descripcion4 = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[3]/div[2]/div/div/div[2]/div"
path_descripcion5 = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div/div"
path_descripcion6 = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[3]/div/div/div[1]/div[2]/div/div/div"
path_descripcion7 = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[2]/div/div/div/div[2]/div"
path_descripcion8 = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[3]/div/div/div[1]/div[2]/div"
path_caracteristicas = ["/html/body/form/div[3]/div[5]/div[4]/div[2]/div[1]/div[1]/div/div[2]/div/div[3]/div/div[","]/span"]

path_imagen = ["/html/body/form/div[3]/div[5]/div[4]/div[1]/div/div/div/div/div[3]/div[3]/div/div[2]/div[2]/div[5]/div/div[","]/img"]
path_agente = "/html/body/form/div[3]/div[5]/div[4]/div[2]/div[2]/div[1]/div[1]/section/div/div/div/div[1]/div[2]/h4/a"
# configuraciones para ignorar los certificados ssl para descargar las imagenes
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

def esperarPorObjeto(navegador_abierto, tiempo, tipoObjeto, identificadorObjeto, nombre):
    """
        Espera que se cargue el objeto de la pagina
        :return
        bool
    """
    escribir_en_log(f"Esperando a que cargue {nombre}", 1)

    try:
        WebDriverWait(navegador_abierto, tiempo).until(
            expected_conditions.presence_of_element_located((tipoObjeto, identificadorObjeto)))
        return True
    except:

        return False
def ejecutar_script_categoria_comercial(navegador):
    script_asu = (
            "var selectElement = document.evaluate(\"/html/body/form/div[3]/div[3]/div[1]/div/div[3]/div/div/div/div[4]/div/div/select\", document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;"
            "var event = new Event('mousedown');"
            "selectElement.dispatchEvent(event);"
            "selectElement.options[7].selected = true;")
    
    try:
        navegador.execute_script(script_asu)
        escribir_en_log(f"Se ejecuto el script para cambiar a estancia/terreno", 1)
    except:
        escribir_en_log(f"No se pudo cambiar la categoria a estancia/terreno", 3)


# esta funcion carga la pagina principal de remax y realiza la busqueda de las propiedades en la ciudad que le pasemos
# solo probe con asuncion
def buscar_ciudad(ciudad, navegador):
    navegador.get(url)
    """script = ("var campo_ciudad = document.getElementById(\"geolocctrl\");"
              "campo_ciudad.value = ''")
    navegador.execute_script(script)"""
    hay_resultados = True
    if ciudad == "Asuncion":
        # Escribir en el campo de busqueda la ciudad
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Asuncion")
        time.sleep(3)
    elif ciudad == "Sanber":
        # Escribir en el campo de busqueda la ciudad
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("San Ber")
        time.sleep(3)
    elif ciudad == "Fernando":
        # Escribir en el campo de busqueda la ciudad
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Fer")
        time.sleep(3)
    elif ciudad == "Sanlo":
        # Escribir en el campo de busqueda la ciudad
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("San Lo")
        time.sleep(3)
    elif ciudad == "Luque":
        # Escribir en el campo de busqueda la ciudad
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Luque")
        time.sleep(3)
    elif ciudad == "Lamba":
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Lamba")
        time.sleep(3)
    elif ciudad == "Altos":
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Altos")
        time.sleep(3)
    elif ciudad == "Aregua":
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Aregua")
        time.sleep(3)
    elif ciudad == "Paraguay":
        # ejecuta un script para seleccionar la categoria que requiere
        ejecutar_script_categoria_comercial(navegador)
        time.sleep(1)
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Paraguay")
        time.sleep(3)
    elif ciudad == "VillaElisa":
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Villa Eli")
        time.sleep(3)
    elif ciudad == "Presidente":
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Presidente")
        time.sleep(3)
    elif ciudad == "Ñemby":
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Ñemby")
        time.sleep(3)
    elif ciudad == "Capiata":
        navegador.find_element(By.XPATH, path_campo_ciudad).send_keys("Capiata")
        time.sleep(3)

    # click para realizar la busqueda
    navegador.find_element(By.XPATH, "/html/body/form/div[3]/div[3]/div[1]/div/div[3]/div/div/div/div[7]/button").click()
    # esperar a que cargue la pagina
    esperarPorObjeto(navegador, 30, By.XPATH,  path_boton_siguiente, "Algun resultado de la pagina")
    ciudad_busqueda_actual = ciudad

def extraer_links(links_por_resultados, navegador, ciudad, cantidad_agregar):

    """
    Abre la pagina de remax y realiza la extraccion de todos los resultados por ciudad
    links_por_resultados : cantidad de resultados por pagina
    """
    escribir_en_log(f"Comienzo la funcion [extraer_links] [ciudad:{ciudad}]", 1)
    links = 0
    global hay_resultados
    contador_iteracion = 1
    escribir_en_log(f"La variable hay resultados indica si todavia hay una pagina siguiente para sacar resultados", 1)
    escribir_en_log(f"Se inicia el bucle para sacar links [hay_resultados:{hay_resultados}]", 1)
    while hay_resultados:
        # los links que se van a abrir en cada iteracion
        escribir_en_log(f"[Iteracion:{contador_iteracion}]", 1)
        contador = 0

        link_extraidos = []
        escribir_en_log("Se validara 2 veces los links para extraer", 1)
        while contador < 2:# dos equivale a la cantidad de intentos o vueltas para sacar los links de cada resultado
            escribir_en_log(f"[Validacion:{contador+1}]", 1)

            for indice in range(1, links_por_resultados):#26 es la cantidad resultados que retorna la pagina para llamar a la funcion por default con esa cantidad

                path_resultado = f"/html/body/div[1]/form/div[3]/div[5]/div/div[8]/div/div[3]/div/div/div[3]/div[2]/div[1]/div[1]/div/div[{indice}]/div/div[6]/span/a"
                path_tarjeta = f"/html/body/div[1]/form/div[3]/div[5]/div/div[8]/div/div[3]/div/div/div[3]/div[2]/div[1]/div[1]/div/div[{indice}]/div/div[3]/div/span"

                try:
                    estado_propiedad = navegador.find_element(By.XPATH, path_tarjeta).text
                    escribir_en_log(f"[Iteracion:{contador_iteracion}][indice:{indice}] [estado_propiedad:{estado_propiedad}]",1)
                except:
                    estado_propiedad = "Sin Estado"
                    escribir_en_log(f"[Iteracion:{contador_iteracion}][indice:{indice}] [estado_propiedad:{estado_propiedad}]", 1)
                # para los casos donde el estado de la propiedad es vendida o reservada

                if estado_propiedad not in tipo_propiedad_excluir:

                    try:
                        if links >= cantidad_agregar:

                            try:
                                base_remax.to_excel(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.xlsx'), index=False)
                                escribir_en_log("Se actualizo el archivo excel", 1)
                            except:
                                escribir_en_log("Error al querer sobrescribir el excel (puede ser que este abierto)", 1)
                                pass
                            hay_resultados = False
                            escribir_en_log(f"se cambia estado de la variable [hay_resultados:{hay_resultados}]", 1)
                            break
                        #asd
                        try:
                            link_resultado = navegador.find_element(By.XPATH, path_resultado).get_attribute("href")
                            escribir_en_log(f"[Iteracion:{contador_iteracion}][indice:{indice}]Se obtiene el link: {link_resultado}", 1)
                        except:
                            escribir_en_log(f"No se pudo sacar el link [indice:{indice}] ", 1)
                            pass
                        try:
                            tipo = navegador.find_element(By.XPATH, f"{path_tipo_propiedad[0]}{indice}{path_tipo_propiedad[1]}").text
                            escribir_en_log(f"[Iteracion:{contador_iteracion}][indice:{indice}]Se saco el tipo de propiedad [tipo:{tipo}]", 1)
                        except:
                            tipo = "Sin Tipo"
                            escribir_en_log(f"No se pudo extraer el tipo de propiedad [tipo:{tipo}]", 3)

                        base_remax = pd.read_csv(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.csv'))
                        validacion_base_link = len(base_remax.loc[base_remax['link'] == link_resultado]) > 0
                        escribir_en_log(f"[Iteracion:{contador_iteracion}][indice:{indice}]Se valida si ya existe el link en la base [validacion_base_link:{validacion_base_link}]", 1)
                        if not validacion_base_link:# si no existe el link en la base
                            nueva_fila = [{'titulo': '', 'tipo': tipo, 'precio': '',
                                           'descripcion': '', 'link': link_resultado, 'ide': '',
                                           'ciudad': ciudad, 'publicado_facebook': '',
                                           'fecha_inserion': datetime.now().strftime("%d/%m/%Y"),
                                           'publicado_clasipar': '', 'publicado_info': '',
                                           'publicado_hendyla': ''}]
                            link_extraidos.append(link_resultado)
                            df_nueva_fila = pd.DataFrame(nueva_fila)
                            links += 1

                            base_remax = pd.concat([base_remax, df_nueva_fila], ignore_index=True)
                            base_remax.to_csv(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.csv'), index=False)
                            escribir_en_log(f"[Iteracion:{contador_iteracion}][indice:{indice}]Se actualizo la base [remax_propiedades.csv] con un nuevo link",1)

                            # Para crear un excel
                            try:
                                base_remax.to_excel(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.xlsx'), index=False)
                                escribir_en_log("Se actualizo el archivo excel", 1)
                            except:
                                escribir_en_log("Error al querer sobrescribir el excel (puede ser que este abierto)", 1)
                                pass

                            if links >= cantidad_agregar:
                                escribir_en_log(f"Se alcanzo la cantidad establecida para agregar {cantidad_agregar}", 1)
                                hay_resultados = False
                                break



                    except:
                        pass
            contador += 1
            if links >= cantidad_agregar:
                escribir_en_log(f"Se alcanzo la cantidad de {cantidad_agregar} resultados", 1)
                hay_resultados = False
                break

        if hay_resultados:
            try:
                esperarPorObjeto(navegador, 10, By.XPATH, path_boton_siguiente, "Algun resultado de la pagina")
                navegador.find_element(By.XPATH, path_boton_siguiente).click()
                escribir_en_log("Se intenta clickear el boton siguiente", 1)
                no_cargaron_nuevos = True
                veces_intentandas = 0# contador para refrescar el navegador en caso que quede pegada a una pestana

                while no_cargaron_nuevos:

                    time.sleep(1)
                    #print("Siguen siendo los mismos resultados")
                    path_resultado_generico = f"/html/body/div[1]/form/div[3]/div[5]/div/div[8]/div/div[3]/div/div/div[3]/div[2]/div[1]/div[1]/div/div[{3}]/div/div[6]/span/a"
                    if navegador.find_element(By.XPATH, path_resultado_generico).get_attribute("href") not in link_extraidos:
                        escribir_en_log(f"Esperando a que carguen los resultados [veces_intentandas:{veces_intentandas}]",
                                        1)
                        time.sleep(1)
                        no_cargaron_nuevos = False
                        veces_intentandas += 1
                        # si supera la espera 5 veces recarga el navegador y vuelve a clickear siguiente
                        if veces_intentandas > 5:
                            escribir_en_log(f"Se llego a la cantidad maxima de intetos y espera [veces_intentadas:{veces_intentandas}]", 1)
                            veces_intentandas = 0
                            escribir_en_log("Se intenta clickear el boton siguiente", 1)
                            navegador.find_element(By.XPATH, path_boton_siguiente).click()


            except Exception as ex:
                escribir_en_log("No se encontro disponible el boton siguiente (Ya se alcanzo el final de los resultados)", 1)

                hay_resultados = False

            contador_iteracion += 1
    hay_resultados = True

# funciones para quitar datos
def extraer_titulo(navegador):
    """
    Extrae el titulo del resultado abierto
    """
    titulo = navegador.find_element(By.XPATH, path_titulo).text
    escribir_en_log(f"Se extrae el titulo:{titulo}", 1)
    return titulo
def extraer_precio(navegador):

    precio = "₲"
    for precios in [path_precio, path_precio2]:

        try:
            precio = navegador.find_element(By.XPATH, precios).text

            break
        except:
            pass


    if "₲" in precio:
        precio = precio.replace("₲", "").replace(",", "").rstrip().lstrip() + " GS"
    if "$" in precio:
        precio = precio.replace("₲", "").replace(",", "").rstrip().lstrip() + " USD"
    escribir_en_log(f"Se extrae el precio y tipo_moneda: {precio}", 1)
    return precio

def extraer_id(navegador, base_remax, link):
    mensaje = ""
    eliminar = False
    try:
        mensaje = navegador.find_element(By.XPATH,
                                         "/html/body/form/div[3]/div[5]/div/div/div/div[1]/div[1]/div[1]/h1").text
        if "Envíenos un mensaje" in mensaje:
            escribir_en_log(f"Ya se elimino la propiedad ", 2)
            eliminar = True

    except:
        try:
            mensaje = navegador.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[4]/div/div/div/h1").text
        except:
            pass
        if "404 PAGINA NO ENCONTRADA" in mensaje:
            eliminar = True
            escribir_en_log(f"Ya se elimino la propiedad de remax ", 2)
    if eliminar:
        base_remax.loc[base_remax['link'] == link, 'descripcion'] = "Eliminado"
        base_remax.loc[base_remax['link'] == link, 'titulo'] = "Eliminado"
        for num in range(1, 7):
            base_remax.loc[base_remax[f"{num}publicado_info"] == link, 'titulo'] = 1
            base_remax.loc[base_remax[f"{num}publicado_clasipar"] == link, 'titulo'] = 1

        base_remax.to_csv(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.csv'), index=False)
        escribir_en_log(f"Se actualiza la base", 1)
        try:
            base_remax.to_excel(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.xlsx'), index=False)
        except:
            pass
    id = "Error"
    escribir_en_log(f"Se intentara sacar el ide entre los indices 3 y 11", 1)
    indice = 3
    for indice in range(3, 11):

        try:
            id = navegador.find_element(By.XPATH, f"{path_id[0]}{indice}{path_id[1]}").text

            if "ID:" in id:
                id = id.replace("ID:", "").rstrip().lstrip()
                escribir_en_log(f"[indice:{indice}]Se encontro el id", 1)
                break

        except:
            escribir_en_log(f"[indice:{indice}]No se pudo extraer el ide", 3)
            pass

    if indice == 10:
        escribir_en_log(f"No se pudo extraer el ide", 3)
        return "Sin ID"
    else:
        escribir_en_log(f"Se extrajo el [ide:{id}]", 1)
        return id
def extraer_atributos(navegador):
    seguir = True
    indice = 1
    text_atributos = ""
    atributos = {}
    escribir_en_log(f"Se extrae tabla de descripcion", 1)
    while seguir:
        try:
            # nombre atributo
            nom = navegador.find_element(By.XPATH, f"{path_atributo[0]}{indice}{path_atributo[1]}").text
            valor = navegador.find_element(By.XPATH, f"{path_valor[0]}{indice}{path_valor[1]}").text
            text_atributos += f"{nom} {valor}\n"
            atributos[nom] = valor

        except:
            seguir = False

        indice += 1
    escribir_en_log(f"Datos tabla: {atributos}", 1)
    return atributos
def extraer_descripcion(navegador):
    contador = 1
    descri = ""
    for desc in [path_descripcion, path_descripcion2, path_descripcion3, path_descripcion4,
                 path_descripcion5, path_descripcion6, path_descripcion7, path_descripcion8]:

        try:
            descri = navegador.find_element(By.XPATH, desc).text
            escribir_en_log(f"Se extrae la descripcion", 1)
            break
        except:
            escribir_en_log(f"[intento:{contador}]Se intento sacar la descripcion", 1)

        contador += 1
    if descri == "":
        escribir_en_log(f"No se pudo extraer la descripcion", 3)

    return descri
def extraer_ruta_imagenes(navegador):
    seguir = True
    indice = 1
    imagenes = []
    while seguir:
        try:
            ruta = navegador.find_element(By.XPATH, f"{path_imagen[0]}{indice}{path_imagen[1]}").get_attribute("src")
            escribir_en_log(f"[link_imagen:{ruta}]", 1)
            imagenes.append(ruta)

            if len(imagenes) > 8:
                escribir_en_log(f"Se llego al maximo de imagenes", 1)
                seguir = False
            indice += 1
        except:

            break
    return imagenes
def crear_carpetas_datos(ide):
    ruta_carpeta_crear = PurePath(RUTA_DATOS, ide)
    carpeta = Path(ruta_carpeta_crear)
    escribir_en_log(f"[ide:{ide}]Se crea la carpeta", 1)
    carpeta.mkdir(parents=True, exist_ok=True)
    carpeta = Path(PurePath(ruta_carpeta_crear, "img"))
    carpeta.mkdir(parents=True, exist_ok=True)

def validar_habitaciones(navegador, base_csv, link):
    atributos = extraer_atributos(navegador)
    mts = False
    for atri in atributos:
        if 'Dormitorios' in atri:
            base_csv.loc[base_csv['link'] == link, 'habitaciones'] = atributos[atri]

        if 'Baños' in atri:
            base_csv.loc[base_csv['link'] == link, 'banio'] = atributos[atri]

        if "Sup. Lote (m²)" in atri:
            base_csv.loc[base_csv['link'] == link, "mts"] = atributos[atri]
            mts = True

        if "Area de Construcción (m²)" in atri:
            base_csv.loc[base_csv['link'] == link, "area"] = atributos[atri]

    if not mts:
        for atri in atributos:
            if "Total Mts²" in atri:
                base_csv.loc[base_csv['link'] == link, "mts"] = atributos[atri]
def extraer_agente_inmobiliario(navegador):
    """
    Extrae el agente inmobiliario correspondiente a la propiedad
    """
    try:
        agente = navegador.find_element(By.XPATH, path_agente).text
        escribir_en_log(f"Agente: {agente}", 1)
    except:
        escribir_en_log(f"No se pudo obtener el agente", 1)
        agente = ""



    agente = agente.rstrip().lstrip()
    return agente

def preparar_texto(tip, ide, navegador):

    base_remax = pd.read_csv(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.csv'))




    archivo_txt = open(PurePath(RUTA_DATOS, ide, "Info.txt"), "w", encoding='utf-8')
    titulo = extraer_titulo(navegador)
    precio = extraer_precio(navegador)
    descripcion = extraer_descripcion(navegador)
    agente = extraer_agente_inmobiliario(navegador)
    archivo_txt.write(f"TITULO\n{titulo}\nTIPO\n{base_remax[base_remax['link'] == tip]['tipo']}\nPRECIO\n{precio}\nDESCRIPCION\n{descripcion}\nLINK\n{tip[0]}")
    archivo_txt.close()
    if descripcion != "":

        base_remax.loc[base_remax['link'] == tip, 'descripcion'] = descripcion
        base_remax.loc[base_remax['link'] == tip, 'titulo'] = titulo
        base_remax.loc[base_remax['link'] == tip, 'precio'] = precio
        base_remax.loc[base_remax['link'] == tip, 'ide'] = ide
        base_remax.loc[base_remax['link'] == tip, 'agente_remax'] = agente
        validar_habitaciones(navegador, base_remax, tip)


        #base_remax.loc[base_remax['link'] == tip, 'ciudad'] = busqueda_ciudad_actual

        base_remax.to_csv(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.csv'), index=False)
        escribir_en_log(f"Se actualiza la base", 1)
        try:
            base_remax.to_excel(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.xlsx'), index=False)
        except:
            pass

def descargar_imagenes(imagenes, ide):
    contador = 1
    escribir_en_log(f"[ide:{ide}]Imagenes {imagenes[0:2]}....", 1)
    # separar el guion
    try:
        ide_para_img = ide.split("-")[0]
    except:
        ide_para_img = 1
    # lista las carpetas dentro de datos
    carpetas_imagenes = os.listdir(RUTA_DATOS)
    # si existe la carpeta
    if ide in carpetas_imagenes:
        escribir_en_log(f"[ide:{ide}]Existe la carpeta con el ide para las imagenes", 1)
        # lista las imagenes dentro de la carpeta
        imagenes_lista = os.listdir(PurePath(RUTA_DATOS, ide, "img"))
        # si la lista de imagenes es igual a 0
        if len(imagenes_lista) == 0:
            escribir_en_log(f"[ide:{ide}]Comienza la descarga de imagenes", 1)
            escribir_en_log(f"[ide:{ide}]Se decargan las imagenes", 1)
            for ruta in imagenes:
                inicio_descarga = time.time()
                with urllib.request.urlopen(ruta, context=ctx) as u, open(PurePath(RUTA_DATOS, ide, "img", f"{contador}_img_{ide_para_img}.jpg"), 'wb') as f:
                    escribir_en_log(f"[ide:{ide}]Se descargo la imagen {ide_para_img}", 1)
                    f.write(u.read())
                    fin_descarga = time.time()
                    duracion_descarga = fin_descarga - inicio_descarga
                    escribir_en_log(f"[ide:{ide}]Tiempo de descarga {duracion_descarga:.2f}", 1)



                contador += 1
        else:
            escribir_en_log(f"[ide:{ide}]Ya tenia imagenes", 2)
def recorrer_links_de_resultados(navegador, cantidad_scrapear):
    escribir_en_log("Se ejecuta la funcion recorrer_links_de_resultados", 1)
    base_remax = pd.read_csv(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.csv'))
    try:
        base_remax.to_excel(PurePath(RUTA_BOT, 'driver', 'remax_propiedades.xlsx'), index=False)
    except:
        pass
    #print("link", links[0])
    contador_resultados_validos_descargados = 0

    contador = 1

    cant = len(base_remax[(base_remax['titulo'].isna()) | (base_remax['titulo'] == '')]['link'].to_list())
    escribir_en_log(f"Pendientes para descarga de imagenes e informacion {cant}", 1)
    contador_proceso = 1
    for link in base_remax[(base_remax['titulo'].isna()) | (base_remax['descripcion'] == '')]['link'].to_list():
        escribir_en_log(f"Procesando {contador_proceso} de {cantidad_scrapear}", 1)

        # validar antes de scrapear
        # valida que el link que se va a recorrer no tenga ya los datos scrapeados
        if len(base_remax.loc[(base_remax['link'] == link) & (base_remax['titulo'] == "")].index) < 1:#valida que no se descargue o scrapee los datos de una propiedad que ya se cargo
            escribir_en_log(f"Se abre el link '{link}'", 1)
            navegador.get(link)

            esperarPorObjeto(navegador, 20, By.XPATH, path_titulo, "El titulo o imagenes de la propiedad")
            ide = extraer_id(navegador, base_remax, link)

            if ide == "Sin ID":
                escribir_en_log(f"Esta propiedad no tiene ide se pasa a la siguiente", 1)
                ide += f" duplicado{contador}"
                contador += 1
                pass
            else:


                try:
                    crear_carpetas_datos(ide)
                    imagenes = extraer_ruta_imagenes(navegador)
                    preparar_texto(link, ide, navegador)
                    descargar_imagenes(imagenes, ide)
                    contador_resultados_validos_descargados += 1
                    contador_proceso += 1
                    if contador_resultados_validos_descargados >= cantidad_scrapear:# cambiar a 60
                        escribir_en_log(f"Se descargaron con exito los datos de {cantidad_scrapear} propiedades", 1)
                        print("\n\n")
                        break



                except :

                    pass
        else:
            escribir_en_log(f"Esta propiedad ya tiene datos", 1)

def variables_configuracion():
    """
    Lee la configuracion establecida en el archivo excel en la carpeta driver
    Pass.xlsx['Configuracion']
    """
    # abre el excel
    excel = openpyxl.load_workbook(PurePath(RUTA_BOT, "Driver", "Pass.xlsx"), read_only=True)
    # se saca la hoja Configuracion
    hoja = excel.worksheets[3]
    # La columna es B
    columna_valor = 2

    cantidad_propiedades = hoja.cell(row=2, column=columna_valor).value
    cantidad_publicar = hoja.cell(row=3, column=columna_valor).value
    cantidad_agregar = hoja.cell(row=4, column=columna_valor).value

    excel.close()
    return {'cantidad_propiedades': cantidad_propiedades,
            'cantidad_publicar': cantidad_publicar,
            'cantidad_agregar': cantidad_agregar}




print(variables_configuracion())