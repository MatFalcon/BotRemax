import openpyxl as op
import pandas as pd
excel = op.load_workbook("MATIII.xlsx")
hoja = excel.worksheets[0]

max_filas = hoja.max_row + 1
max_cols = hoja.max_column + 1
ignorar = ["Guaranies", "Obs. OC:", "Obs. Prob.:"]
terminar = "Filtros Establecidos"
cols = {
    "Nro Oc": [],
    "Fecha": [],
    "F.Entrega": [],
    "Sucursa": [],
    "Proveedor": [],
    "Importe": [],
    "Estado": [],
    "F.Recep": [],
    "Usuario": [],
    "Edi": [],
    "Aut":[]
        }
for fila in range(7, max_filas):
    celda = hoja.cell(fila, 1).value
    if celda not in ignorar and celda is not None:
        cols["Nro Oc"].append(celda)
        cols["Fecha"].append(hoja.cell(fila, 3).value)
        cols["F.Entrega"].append(hoja.cell(fila, 9).value)
        cols["Sucursa"].append(hoja.cell(fila, 12).value)
        cols["Proveedor"].append(hoja.cell(fila, 15).value)
        cols["Importe"].append(hoja.cell(fila, 18).value)
    if celda == terminar:
        break
df = pd.DataFrame(cols)
df.to_excel("Res.xlsx")
print(cols)
excel.close()